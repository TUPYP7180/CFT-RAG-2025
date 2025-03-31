
import openai
from langsmith import traceable
from langsmith.wrappers import wrap_openai

import os

from volcenginesdkarkruntime import Ark

client = Ark(base_url="https://ark.cn-beijing.volces.com/api/v3")


from dotenv import load_dotenv
from src.build_index import load_vec_db
from src.rag_complete import augment_prompt
from trag_tree import build, hash
import time
import argparse
import openpyxl

from langsmith.evaluation import LangChainStringEvaluator, evaluate
from langsmith.schemas import Example, Run

load_dotenv()



vec_db_key = "院志"
entities_file_name = "entities_file"
node_num_max = 2000000

evalutorType = "custom_score_match"
dataset_name = "t-rag-new"

class RagBot:
    def __init__(self, tree_num_max: int, search_method: int, model: str = "ep-20241101151030-lfwks"):
        self._model = model
        self._search_method = search_method  # 保存当前的 search_method

        # 初始化客户端，这里假设 client 已经在其他地方定义
        self._client = client

        # 加载向量数据库
        self.vec_db = load_vec_db(vec_db_key, "vec_db_cache/")
        print("Load vector db finished")

        # 构建森林和 NLP 模块
        self.forest, self.nlp = build.build_forest(tree_num_max, entities_file_name, search_method, node_num_max)
        print("Load forest and nlp finished")

        # 根据 search_method 执行不同的初始化
        if search_method in [4]:
            for entity_tree in self.forest:
                entity_tree.bfs_hash()

        if search_method in [7]:
            hash.cuckoo_build(tree_num_max, node_num_max)

    def retrieve_docs(self, question):
        return augment_prompt(
            question,
            self.vec_db,
            self.forest,
            self.nlp,
            search_method=self._search_method,
            model_name=self._model,
            debug=True
        ),

    def get_answer(self, question: str):
        similar = self.retrieve_docs(question)
        response = self._client.chat.completions.create(
            model=self._model,
            messages=[
                {
                    "role": "system",
                    "content": "你是一个了解医院历史信息的资深专家，接下来我会告诉你一些信息，用中文回答",
                },
                {"role": "user", "content": question},
            ],
        )
        return {
            "answer": response.choices[0].message.content,
            "contexts": [str(doc) for doc in similar],
        }

# 定义参数组合
parameter_combinations = [
    {'tree_num_max': 50, 'search_method': 1},
    {'tree_num_max': 50, 'search_method': 2},
    {'tree_num_max': 50, 'search_method': 5},
    {'tree_num_max': 50, 'search_method': 7},
    {'tree_num_max': 300, 'search_method': 0},
    {'tree_num_max': 300, 'search_method': 1},
    {'tree_num_max': 300, 'search_method': 2},
    {'tree_num_max': 300, 'search_method': 5},
    {'tree_num_max': 300, 'search_method': 7},
    {'tree_num_max': 600, 'search_method': 0},
    {'tree_num_max': 600, 'search_method': 1},
    {'tree_num_max': 600, 'search_method': 2},
    {'tree_num_max': 600, 'search_method': 5},
    {'tree_num_max': 600, 'search_method': 7},
]

repeat = 20
scoresum = 0

# 定义自定义评分函数
def custom_score_match(run: Run, example: Example) -> dict:
    reference = example.outputs["answer"]
    prediction = run.outputs["answer"]

    # 调用评估模型获取评分
    completion = client.chat.completions.create(
        model="ep-20241101151030-lfwks",
        messages = [
            {"role": "system", "content": "你是一个回答评估专家，给定两个字符串prediction和reference，你需要给prediction和reference之间的相似程度打分，分数为0-100之间的整数，只准给出一个阿拉伯数字，不准输出任何其他额外内容，尽可能在0-100之间的区分度比较大，不要集中在某一个区间"},
            {"role": "user", "content": f"prediction:{prediction.lower()}, reference: {reference.lower()}"},
        ],
    )
    score = int(completion.choices[0].message.content)
    print(f"score: {score}")
    global scoresum
    scoresum = scoresum + score
    print(f"scoresum: {scoresum}")
    
    return {"key": "custom_score_match", "score": score}




# 创建或加载 Excel 文件
file_path = "evaluation_results.xlsx"
try:
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # 初始化表头
    sheet.append(["tree_num", "method"] + [f"score-{i}" for i in range(1, 21)] + ["score-avg"])
    workbook.save(file_path)



# 运行测试
def run_tests():
    row_index = sheet.max_row + 1  # 从当前空行开始写
    for combination in parameter_combinations:
        tree_num_max = combination['tree_num_max']
        search_method = combination['search_method']

        print(f"当前参数组合: tree_num_max={tree_num_max}, search_method={search_method}")
        rag_bot = RagBot(tree_num_max=tree_num_max, search_method=search_method)

        def predict_rag_answer(example: dict):
            """Use this for answer evaluation"""
            response = rag_bot.get_answer(example["prompt"])
            return {"answer": response["answer"]}

        # 初始化结果记录
        scores = []

        for i in range(1, repeat + 1):
            global scoresum
            scoresum = 0
            command_str = f"tree_num_max={tree_num_max}, search_method={search_method}, run={i}"
            print(f"正在执行: {command_str} (第 {i} 次)")

            try:
                experiment_results = evaluate(
                    predict_rag_answer,
                    data=dataset_name,
                    evaluators=[custom_score_match],
                    experiment_prefix=f"{dataset_name}-{evalutorType}-tree-num-{tree_num_max}-method-{search_method}",
                    metadata={"variant": "t-rag"},
                )
                # 记录 scoresum/10 的值
                average_score = round(scoresum / 10, 3)
                scores.append(average_score)
                print(f"执行完成: {command_str} (第 {i} 次), 平均分: {average_score}")
            except Exception as e:
                print(f"命令失败: {command_str} (第 {i} 次), 错误: {e}")
                scores.append(0)  # 记录错误时的默认值
            print("----------------------------------------")
        
        # 计算整体平均分
        overall_avg = round(sum(scores) / len(scores), 3)
        print(f"参数组合完成: tree_num_max={tree_num_max}, search_method={search_method}, 平均分: {overall_avg}")

        # 写入 Excel 文件
        row  = [tree_num_max, search_method] + scores + [overall_avg]
        print(f"row: {row}")
        sheet.append(row)
        workbook.save(file_path)  # 每次运行后保存文件
        
        print("========================================")

if __name__ == "__main__":
    run_tests()
